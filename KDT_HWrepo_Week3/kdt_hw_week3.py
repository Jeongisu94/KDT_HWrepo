from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import smtplib
import re #regular expression

SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
SMTP_USER = ''
SMTP_PASSWORD = ''

def send_mail(name, addr, subject, contents, attachment=None):
    if not re.match('(^[a-zA-Z0-9_.-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        print('Wrong email')
        return

    msg = MIMEMultipart('alternative')
    if attachment:
        msg = MIMEMultipart('mixed')

    msg['From'] = SMTP_USER
    msg['To'] = addr
    msg['Subject'] = name + '님, ' + subject

    text = MIMEText(contents, _charset='utf-8')
    msg.attach(text)

    if attachment:
        from email.mime.base import MIMEBase
        from email import encoders

        file_data = MIMEBase('application', 'octect-stream')
        file_data.set_payload(open(attachment, 'rb').read())
        encoders.encode_base64(file_data)

        import os
        filename = os.path.basename(attachment)
        file_data.add_header('Content-Disposition', 'attachment; filename="' + filename + '"')
        msg.attach(file_data)

    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    smtp.close()

#### 프로젝트 폴더에 있는 email_list.xlsx 파일에 이메일 받을 사람들의 정보를 입력하세요.
#### 엑셀 파일의 정보를 읽어올 수 있는 모듈을 import하세요.
#### email_list.xlsx 파일을 읽어와 해당 사람들에게 수집한 뉴스 정보 엑셀 파일을 send_mail 함수를 이용해 전송하세요.

from openpyxl import load_workbook
wb = load_workbook('email_list.xlsx', read_only=True)
data = wb.active
receivers = []

#email_list의 파일에서 이름과 이메일을 가져와 리스트에 저장
for names, emails  in zip (data.iter_rows(min_row=3,min_col=2, max_col=2),data.iter_rows(min_row=3,min_col=3,max_col=3)):
    for name, email in zip(names, emails):
        tempReceive = [name.value,email.value]
        receivers.append(tempReceive)


#from NaverNewsCrawler import NaverNewsCrawler
#crawler = NaverNewsCrawler("패스트 캠퍼스") # 네이버뉴스에서 데이터 크롤링
#crawler.get_news("data.xlsx")

crawledWorkbook = load_workbook('data.xlsx', read_only=True)
crawledData = crawledWorkbook.active
contents = ''

for row in crawledData.iter_rows():
    for cell in row:
        contents = contents + " "+ str(cell.value)
    contents += '\n'

for name , mail in receivers:
    send_mail(name,mail,'Naver News : ',contents)
